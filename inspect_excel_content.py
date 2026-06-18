import openpyxl
from pathlib import Path

excel_path = Path("excel_sheets/bhagavata_book2_ch1_linear.xlsx")
if excel_path.exists():
    wb = openpyxl.load_workbook(excel_path)
    print("Sheets in linear excel:", wb.sheetnames)
    ws = wb.active
    print("Linear Sheet dimensions:", ws.dimensions)
    print("First 10 rows:")
    for r in range(1, 15):
        row_vals = [ws.cell(r, c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")
else:
    print("File not found.")
