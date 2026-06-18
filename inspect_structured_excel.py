import openpyxl
from pathlib import Path

excel_path = Path("structured/bhagavata_book2_ch1_structured.xlsx")
if excel_path.exists():
    wb = openpyxl.load_workbook(excel_path)
    print("Sheets in structured excel:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\nSheet: {name}, dimensions: {ws.dimensions}")
        print("First 3 rows:")
        for r in range(1, 4):
            row_vals = [ws.cell(r, c).value for c in range(1, 15)]
            print(f"Row {r}: {row_vals}")
else:
    print("File not found.")
